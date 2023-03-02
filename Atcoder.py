import os
import csv
import requests
from bs4 import BeautifulSoup
import re
import codecs
import openpyxl as excel
import random


# URLのクラス
class Url():
    def __init__(self, url):
        self.url = url
    
    # urlのページのhtmlからsection項目を抽出する
    def get_section(self, lang = 'all'):
        r = requests.get(self.url)
        soup = BeautifulSoup(r.content, "html.parser")
        temp = str(soup.select("section"))
        temp = temp.replace('[', '')
        temp = temp.replace(']', '')
        if lang == 'all':
            return temp
        elif lang == 'japanese':
            temp = temp.split(', <section>\n<h3>Problem Statement')[0]
            return temp
        elif lang == 'english':
            temp = temp.split(', <section>\n<h3>Problem Statement')[1]
            temp = '<section>\n<h3>Problem Statement' + temp
            return temp
    
    # 問題番号を取得する
    def get_problem_number(self):
        num = re.search(r'\d\d\d', self.url).group()
        return num
    
    # abc???_?の形で問題番号と記号を取得する
    def get_index(self):
        return self.url[-8:]
    
    # 記号を取得する
    def get_alphabet(self):
        return self.url[-1]

    # 日本語のsectionだけを取ってくる
    def get_japanese_section(self):
        a = self.get_section()
        a = a.split(', <section>\n<h3>Problem Statement')[0]
        return a
    
    # 英語のsectionだけを取ってくる
    def get_english_section(self):
        a = self.get_section()
        a = a.split(', <section>\n<h3>Problem Statement')[1]
        a = '<section>\n<h3>Problem Statement' + a
        return a

# 文章を扱うクラス
class Text():
    original = None
    corrected = None

    def __init__(self, text):
        self.original = text
        self.corrected = text
        # タグを削除する
        delete_word = ["<var>", "</var>", "<code>", "</code>", "<strong>", "</strong>", "<li>", "</li>", "<ul>", "</ul>", '<details>', '</details>', '<summary>', '</summary>', '<blockquote>', '</blockquote>']
        for i in delete_word:
            self.corrected = self.corrected.replace(i, '')
        # いったん保留（2023/1/15)
        # 不等号（\leを<=）などを置き換える。≤は環境依存文字だったので避けた
        # 置き換え前 = ['\le', '\dots', '\\times', '\\vdots', '\sqrt']
        # 置き換え後 = ['<=', '...', '×', '︙', '√']
        # for i in range(len(置き換え前)):
        #     self.corrected = self.corrected.replace(置き換え前[i], 置き換え後[i])
    
    # すべての項目名をリストとして出力する
    def get_items(self):
        items = re.findall(r'<h3>.+?</h3>', self.corrected)
        delete_word = ["<h3>", "</h3>"]
        for i in delete_word:
            for j in range(len(items)):
                items[j] = items[j].replace(i, "")
        return items

    # key(項目名)の内容を取得する
    def extract(self, key, *args):
        flg = -1

        if key == '問題文1':
            flg = 0
            key = '問題文'
        
        elif key == '問題文2':
            flg = 1
            key = '問題文'
        
        elif key == 'Problem Statement 1':
            flg = 0
            key = 'Problem Statement'
        
        elif key == 'Problem Statement 2':
            flg = 1
            key = 'Problem Statement'

        # keyのセクションの始めから終わりまでをを抽出する
        keyword = '<section>\s<h3>' + key + '(.|\s)*?</section>'
        if re.search(keyword, self.corrected) != None:
            txt = re.search(keyword, self.corrected).group()
        else:
            return None
        
        # 抽出した文章の始めにある<section>タグや項目名（問題文、制約など）を消す
        temp = key + '</h3>'
        delete_word = ['<section>\n<h3>', '\n</section>', temp]
        for i in delete_word:
            txt = txt.replace(i, "")

        # 問題文1などの場合、ここで分割する
        if flg == 0 or flg == 1:
            a = re.findall('<p>((.|\s)+?)</p>', txt)
            txt = a[flg][0]
        
        # argsに値があればタグを残したものを出力する
        if len(args) != 0 and args[0] == 0:
            return txt
        
        # 改行するタグを\nに置き換えた後、\nの連続をなくす
        br_word = ['</h3>', '<p>', '</p>', '<br/>', '<pre>', '</pre>','\r']
        for i in br_word:
            txt = txt.replace(i, '\n')
        txt = re.sub('[\n]{2,}', '\n', txt)
        
        # 始めと終わりの改行を削除
        if txt[0:1] == '\n':
            txt = txt[1:]
        if txt[-1:] == "\n":
            txt = txt[:-1]

        return txt


# 問題のURL集合
class UrlSet:
    url = 'https://atcoder.jp/contests/'
    url_list = []
    start = None
    end = None

    # startからendまでの問題文のURL集合をurl_listに入れる
    def collect(self, start, end):
        self.url_list = []
        self.start = start
        self.end = end
        for i in range(start, end+1):
            if i < 20:
                for j in range(1, 5):
                    temp = self.url + 'abc' + str(i).zfill(3) + '/tasks/abc' + str(i).zfill(3) + '_' + str(j)
                    self.url_list.append(temp)
            elif 20 <= i <= 125:
                for j in ['a', 'b', 'c', 'd']:
                    temp = self.url + 'abc' + str(i).zfill(3) + '/tasks/abc' + str(i).zfill(3) + '_' + j
                    self.url_list.append(temp)
            elif 125 < i <= 211:
                for j in ['a', 'b', 'c', 'd', 'e', 'f']:
                    temp = self.url + 'abc' + str(i).zfill(3) + '/tasks/abc' + str(i).zfill(3) + '_' + j
                    self.url_list.append(temp)
            elif 211 < i:
                for j in ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h']:
                    temp = self.url + 'abc' + str(i).zfill(3) + '/tasks/abc' + str(i).zfill(3) + '_' + j
                    self.url_list.append(temp)

# タイトル集合
class TitleSet:
    title_set = set()
    start = None
    end = None

    # startからendまでの問題文のタイトル集合をtitle_setに入れる
    def collect(self, start, end, lang = 'all'):
        title_list = []
        self.start = start
        self.end = end

        urlset = UrlSet()
        urlset.collect(start, end)
        
        for i in urlset.url_list:
            item_list = []
            url = Url(i)
            problem = Text(url.get_section(lang))
            item_list = problem.get_items()
            item_set = set(item_list)
            self.title_set.update(item_set)

class CsvFile:
    path = None
    start = None
    end = None

    def __init__(self, path):
        self.path = path
    
    def write_problems(self, start, end):
        title_list = ['問題番号', '記号', '問題文', '注記', '注意点', '注釈', '制約', '入出力', '入力', '出力', '入出力例', '入力例 1', '出力例 1', '入力例 2', '出力例 2', '入力例 3', '出力例 3', '入力例 4', '出力例 4', '入力例 5', '出力例 5', '入力例 6', '出力例 6']
        self.start = start
        self.end = end
        with open(self.path, mode='w',encoding='utf_8', newline= "") as f:
            writer = csv.writer(f)
            writer.writerow(title_list)
        
        a = UrlSet()
        a.collect(start, end)

        f = open(self.path, mode = 'a', encoding='utf_8', newline= "")
        writer = csv.writer(f)
        for i in a.url_list:
            temp_list = []
            b = Url(i)
            temp_list.append(b.get_problem_number())
            temp_list.append(b.get_alphabet())
            c = Text(b.get_section('japanese'))
            for j in title_list[2:]:
                d = c.extract(j)
                if d != None:
                    temp_list.append(d)
                else:
                    temp_list.append('None')
            writer.writerow(temp_list)
        
        f.close()

# 問題がすでに書かれたCSVファイルのクラス
class 問題csv:
    csv_path = None
    csvの中身 = None
    問題番号リスト = []
    問題番号と記号 = []
    問題項目 = None
    
    def csvの読み込み(self, path):
        self.csv_path = path
        with open(path, encoding='UTF8') as f:
            reader = csv.reader(f)
            l = [row for row in reader]
            self.csvの中身 = l
        
        temp_set = set()

        for i in l[1:]:
            temp_set.add(i[0])
        temp_list = list(temp_set)
        self.問題番号リスト = sorted(temp_list)

        for i in l[1:]:
            self.問題番号と記号.append([i[0], i[1]])
        
        self.問題項目 = self.csvの中身[0]

    def 一問書き込み(self, 問題番号, 記号, フォルダ名 = None):
        対象問題 = None
        for i in self.csvの中身[1:]:
            if i[0] == 問題番号 and i[1] == 記号:
                対象問題 = i
        
        if 対象問題 == None:
            print('問題番号と記号を満たすものはありませんでした。')
        
        問題文 = 対象問題[2]
        分割問題 = 問題文.split('\n')

        項目対問題 = []
        for i in range(3, len(対象問題)):
            if 対象問題[i] != 'None':
                項目対問題.append([self.問題項目[i], 対象問題[i]])

        wbname = 問題番号 + '-' + 記号 + '.xlsx'
        if フォルダ名 != None:
            wbname = フォルダ名 + '/' + wbname
        
        wb = excel.Workbook()
        ws = wb.active

        ws['A1'] = 'カテゴリ'
        ws['B1'] = '文など'

        for i in range(len(分割問題)):
            ws.cell(row = i + 2, column = 1).value = '問題文'
            ws.cell(row = i + 2, column = 2).value = 分割問題[i]

        最終行 = 1 + len(分割問題)

        for i in range(len(項目対問題)):
            ws.cell(row = 最終行 + i + 1, column = 1).value = 項目対問題[i][0]
            ws.cell(row = 最終行 + i + 1, column = 2).value = 項目対問題[i][1]

        wb.save(wbname)
    
    def 全問書き込み(self):
        フォルダ名 = '全問出力' + self.問題番号リスト[0] + '_' + self.問題番号リスト[-1]
        os.makedirs(フォルダ名, exist_ok=True)
        for i in self.問題番号と記号:
            self.一問書き込み(i[0], i[1], フォルダ名)

    def 一問ランダム並び替え出力(self, 問題番号, 記号, フォルダ名 = None, ランダム = 1, 乱数シード = 0):
        対象問題 = None
        for i in self.csvの中身[1:]:
            if i[0] == 問題番号 and i[1] == 記号:
                対象問題 = i
        
        if 対象問題 == None:
            print('問題番号と記号を満たすものはありませんでした。')
        
        問題文 = 対象問題[2]
        分割問題 = 問題文.split('\n')

        if ランダム == 1:
            random.seed(乱数シード)
            random.shuffle(分割問題)

        入出力 = []
        for i in range(len(対象問題)):
            if (self.問題項目[i] == '入出力' or self.問題項目[i] == '入力' or self.問題項目[i] == '出力') and 対象問題[i] != 'None':
                入出力.append([self.問題項目[i], 対象問題[i]])
        
        ファイル名 = 問題番号 + '-' + 記号 + '-' + str(乱数シード) + '.py'
        if ランダム != 1:
            ファイル名 = 問題番号 + '-' + 記号 + '.py'

        if フォルダ名 != None:
            ファイル名 = フォルダ名 + '/' + ファイル名
        
        出力文 = ['# 問題文\n']
        for i in 分割問題:
            出力文.append('# ' + i + '\n')
        
        for i in 入出力:
            出力文.append('# ' + i[0] + '\n')
            入出力_改行で分割 = i[1].split('\n')
            for j in 入出力_改行で分割:
                出力文.append('# ' + j + '\n')
        
        出力文.append('def')
        
        with open(ファイル名, 'w', encoding='utf-8') as f:
            f.writelines(出力文)
    
    def 全問ランダム並び替え出力(self):
        フォルダ名 = 'C:/Users/Kotaro/Documents/全問ランダム出力' + self.問題番号リスト[0] + '_' + self.問題番号リスト[-1]
        os.makedirs(フォルダ名, exist_ok=True)
        for i in self.問題番号と記号:
            for j in range(10):
                self.一問ランダム並び替え出力(i[0], i[1], フォルダ名, 乱数シード=j)

    def 全問並び替えずに出力(self):
        フォルダ名 = 'C:/Users/Kotaro/Documents/全問並び替えずに出力' + self.問題番号リスト[0] + '_' + self.問題番号リスト[-1]
        os.makedirs(フォルダ名, exist_ok=True)
        for i in self.問題番号と記号:
            self.一問ランダム並び替え出力(i[0], i[1], フォルダ名, ランダム=0)

def write_txt(txt):
    f = codecs.open('./txt/test3.txt', 'w', encoding ='utf-8')
    f.write(txt)
    f.close()

# a = TitleSet()
# a.collect(276, 276, 'japanese')
# print(a.title_set)

# a = Url('https://atcoder.jp/contests/abc279/tasks/abc279_c')
# a = a.get_section('japanese')
# b = Text(a)
# print(b.extract('制約'))

# a = CsvFile('ABC276-276.csv')
# a.write_problems(276, 276)

a = 問題csv()
a.csvの読み込み('ABC276-276.csv')
a.全問ランダム並び替え出力()